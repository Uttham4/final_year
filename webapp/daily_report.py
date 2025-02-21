from dotenv import load_dotenv
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from backend.backend.API.utils import *
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from msoffcrypto.format.ooxml import OOXMLFile

# Load environment variables
load_dotenv()

# Email configuration
GMAIL = os.getenv("GMAIL", "utthamsing.k.clusterx@gmail.com")
GMAIL_PWD = os.getenv("GMAIL_PWD", "woah igea kqht nvdt")
RECEIVER_GMAIL = os.getenv("RECEIVER_GMAIL", "utthamsing.k.2021.ece@ritchennai.edu.in")


def success_email(subject, body, protected_file_path=None):
    """Send an email with the specified subject, body, optional file attachment."""
    sender_email = GMAIL
    sender_password = GMAIL_PWD
    recipient_email = RECEIVER_GMAIL

    # Set up the email message
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = subject

    # Add email body
    msg.attach(MIMEText(body, 'plain'))

    # Attach the protected file if it exists
    if protected_file_path and os.path.exists(protected_file_path):
        with open(protected_file_path, 'rb') as attachment:
            mime_base = MIMEBase('application', 'octet-stream')
            mime_base.set_payload(attachment.read())
            encoders.encode_base64(mime_base)
            mime_base.add_header(
                'Content-Disposition',
                f'attachment; filename={os.path.basename(protected_file_path)}'
            )
            msg.attach(mime_base)

    try:
        # Connect to Gmail's SMTP server
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        print("Connected to the email server.")

        # Send the email
        server.sendmail(sender_email, recipient_email, msg.as_string())
        print("Email sent successfully.")

    except Exception as e:
        print(f"Error sending email: {e}")

    finally:
        server.quit()



# @retry(Exception, tries=3, delay=10)

# Function to create a password-protected Excel file


def create_excel_report(data_list, count, excel_path, current_date):
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Details"

    # Define styles
    header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Add title and date
    ws.merge_cells('A1:J1')
    ws['A1'] = "Student Details Report"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Generated on: {current_date}"
    ws['A2'].alignment = Alignment(horizontal="center")

    # Add headers
    headers = [
        "First Name", "Last Name", "Roll Number", "Reg Number", 
        "Semester", "Department", "Section", "Address", 
        "Email", "Contact Number"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Add data
    row = 5
    for data in data_list:
        row_data = [
            data['first_name'],
            data['last_name'],
            data['roll_num'],
            data['reg_num'],
            data['semester'],
            data['department'],
            data['section'],
            data['address'],
            data['email'],
            data['contact_number']
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = str(value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
        
        row += 1

    # Add total count row
    row += 1
    ws.cell(row=row, column=1).value = "Total Students"
    ws.cell(row=row, column=10).value = count
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save the workbook
    wb.save(excel_path)

def create_password_protected_excel(data, count, password, output_file):
    # Create an intermediate Excel file
    current_date = datetime.now().strftime("%Y-%m-%d")
    intermediate_excel_path = "temp_student_details.xlsx"
    
    try:
        # Create the formatted Excel file
        create_excel_report(data, count, intermediate_excel_path, current_date)

        # Encrypt the intermediate file and save as password-protected file
        with open(intermediate_excel_path, "rb") as plain_file:
            ooxml_file = OOXMLFile(plain_file)
            with open(output_file, "wb") as encrypted_file:
                ooxml_file.encrypt(password, encrypted_file)
        
        print(f"Password-protected Excel file saved as {output_file}")
        
    except Exception as e:
        print(f"Error creating protected Excel: {e}")
        
    finally:
        # Clean up the intermediate file
        import os
        if os.path.exists(intermediate_excel_path):
            os.remove(intermediate_excel_path)

    
# Main script
def main():
    with DatabaseConnection() as (conn, cursor):
        # Query to fetch data
        query = """
        SELECT first_name, last_name, roll_num, reg_num, semester, 
               department, section, address, email, contact_number 
        FROM student_details 
        WHERE today = True;
        """
        cursor.execute(query)
        data = cursor.fetchall()
        print(data)
        # Query to count rows
        cursor.execute("SELECT COUNT(*) FROM student_details WHERE today = True;")
        # print(cursor.fetchone())
        count = cursor.fetchone()['COUNT(*)']
        print("count",count)
        # File output and password
        output_file = "student_details.xlsx"
        password = "Uttham123"

        # Create the Excel file
        create_password_protected_excel(data, count, password, output_file)
        today_date = datetime.now().strftime("%Y-%m-%d")
    
        # Define email subject
        subject = f"Hostel Food Report - {today_date}"
        body=(
            f"Hello,\n\n"
            f"Please find attached the report for today's hostel food details ({today_date}). "
            f"This report includes information about students and their associated details for today.\n\n"
            f"The file is password-protected for security reasons. Please use the shared password to access the file.\n\n"
            f"Best regards,\n"
            f"Your Hostel Management Team"
        )

        success_email(subject, body, protected_file_path="student_details.xlsx")

if __name__ == "__main__":
    main()
